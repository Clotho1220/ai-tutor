# -*- coding: utf-8 -*-
"""把 Gogo Loves English 1 的單元寫進 Lesson data.xlsx 的「Book 1」工作表。

沿用 Book 3 的排版：單元標題列 → 中文說明列 → 表頭 → Theme/Sentence Pattern/Vocabulary。
重跑本檔會整個覆蓋 Book 1 工作表，內容以此檔為準。
"""
import io
import sys

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

SRC = "Lesson data.xlsx"   # 在專案資料夾內執行
HEADER = ["type", "english", "chinese", "example"]

UNITS = [
    (1, "Hello!", "本單元學習打招呼、自我介紹與說再見。",
     ("Students greet their friends and introduce themselves by name.", "學生們互相問候並自我介紹姓名。"),
     [("What's your name? / I'm [Name].", "詢問對方姓名並自我介紹。", "1. What's your name? 2. I'm Tony."),
      ("Nice to meet you.", "初次見面時的禮貌問候。", "1. Nice to meet you, Tony. 2. Nice to meet you, Jenny.")],
     [("hello (int.)", "哈囉／你好", "Hello, Gogo!"),
      ("goodbye (int.)", "再見", "Goodbye, Tony."),
      ("apple (n.)", "蘋果", "This is an apple."),
      ("ant (n.)", "螞蟻", "Look at the ant.")]),

    (2, "What's this?", "本單元學習詢問並回答教室裡的物品名稱。",
     ("Students ask and answer about various objects found in the classroom.", "學生們詢問並回答教室裡的各種物品。"),
     [("What's this? / It's a/an [object].", "詢問近處不認識的單數物品。", "1. What's this? It's a desk. 2. What's this? It's an eraser.")],
     [("desk (n.)", "書桌", "It's a desk."),
      ("eraser (n.)", "橡皮擦", "It's an eraser."),
      ("book (n.)", "書本", "I have a book."),
      ("pencil (n.)", "鉛筆", "Show me your pencil.")]),

    (3, "Can you sing?", "本單元學習表達自己的能力，並詢問他人會做什麼。",
     ("Students talk about their abilities and ask others what they can do.", "學生們討論自己的能力並詢問他人能做些什麼。"),
     [("Can you [action]? / Yes, I can. / No, I can't.", "詢問對方是否具備某種動作或能力。", "1. Can you read? Yes, I can. 2. Can you fly? No, I can't.")],
     [("sing (v.)", "唱歌", "I can sing."),
      ("read (v.)", "閱讀", "Can you read?"),
      ("draw (v.)", "畫畫", "She can draw."),
      ("cook (v.)", "烹飪", "My dad can cook.")]),

    (5, "Who's she?", "本單元學習介紹與詢問家庭成員。",
     ("Students identify and ask about different members of their family.", "學生們確認並詢問家中的不同成員。"),
     [("Who's she/he? / She's/He's my [family member].", "詢問某位女性或男性是誰。", "1. Who's she? She's my mother. 2. Who's he? He's my father.")],
     [("mother (n.)", "母親／媽媽", "She's my mother."),
      ("father (n.)", "父親／爸爸", "He's my father."),
      ("brother (n.)", "兄弟／哥哥或弟弟", "He's my brother."),
      ("sister (n.)", "姊妹／姊姊或妹妹", "She's my sister.")]),

    (6, "What's his name?", "本單元學習詢問並說出第三人的名字。",
     ("Students ask and state the names of other boys, girls, or teachers.", "學生們詢問並說出其他男孩、女孩或老師的名字。"),
     [("What's his/her name? / His/Her name's [Name].", "詢問第三人（男生／女生）的名字。", "1. What's his name? His name's Ben. 2. What's her name? Her name's Ms. Black.")],
     [("teacher (n.)", "老師", "She's my teacher."),
      ("student (n.)", "學生", "He is a student."),
      ("boy (n.)", "男孩", "The boy is Ben."),
      ("girl (n.)", "女孩", "The girl is Lisa.")]),

    (7, "Is this a dog?", "本單元學習用是非問句辨認動物。",
     ("Students ask questions to identify various animals using yes/no questions.", "學生們使用是非問句來辨認不同的動物。"),
     [("Is this a/an [animal]? / Yes, it is. / No, it isn't.", "詢問這是否為某種特定的動物或物品。", "1. Is this a giraffe? Yes, it is. 2. Is this a whale? No, it isn't.")],
     [("giraffe (n.)", "長頸鹿", "Look at the giraffe."),
      ("whale (n.)", "鯨魚", "A whale is very big."),
      ("penguin (n.)", "企鵝", "The penguin is cute."),
      ("polar bear (n.)", "北極熊", "The polar bear is white.")]),

    (9, "What color is this?", "本單元學習描述並詢問物品的顏色。",
     ("Students describe the color of various items and ask what color they are.", "學生們描述各種物品的顏色，並詢問它們是什麼顏色。"),
     [("What color is this? / It's [color].", "詢問某物品的顏色。", "1. What color is this? It's orange. 2. What color is this? It's red.")],
     [("red (adj.)", "紅色／紅色的", "This apple is red."),
      ("blue (adj.)", "藍色／藍色的", "The sky is blue."),
      ("yellow (adj.)", "黃色／黃色的", "Bananas are yellow."),
      ("green (adj.)", "綠色／綠色的", "The grass is green.")]),

    (10, "How old are you?", "本單元學習詢問與回答年齡，並認識生日相關的單字。",
     ("Students ask and answer about age and celebrate birthdays.", "學生們詢問並回答年齡，以及慶祝生日。"),
     [("How old are you / is [someone]? / I'm [age]. / He's/She's [age].", "詢問對方或第三人的年齡。", "1. How old are you? I'm eight. 2. How old is your brother? He's seven.")],
     [("present (n.)", "禮物", "I have a birthday present."),
      ("candle (n.)", "蠟燭", "Blow out the candle."),
      ("balloon (n.)", "氣球", "The balloon is red."),
      ("cookie (n.)", "餅乾", "I like cookies.")]),

    (11, "What do you like?", "本單元學習談論喜歡的食物與喜好。",
     ("Students talk about their favorite foods and ask friends about their likes and dislikes.", "學生們討論喜愛的食物，並詢問朋友的喜好與厭惡。"),
     [("What do you like? / I like [food].", "詢問對方喜愛什麼食物或事物。", "1. What do you like, Tony? I like cheese. 2. What do you like, Jenny? I like cake.")],
     [("cake (n.)", "蛋糕", "I like chocolate cake."),
      ("cheese (n.)", "起司／乾酪", "Mice like cheese."),
      ("bread (n.)", "麵包", "This bread is fresh."),
      ("salad (n.)", "沙拉", "She eats salad.")]),
]


def main():
    wb = openpyxl.load_workbook(SRC)
    if "Book 1" in wb.sheetnames:
        del wb["Book 1"]
    ws = wb.create_sheet("Book 1", 0)

    for num, title, desc, theme, patterns, words in UNITS:
        ws.append([f"🌟 Unit {num}: {title}"])
        ws.append([])
        ws.append([desc])
        ws.append([])
        ws.append(HEADER)
        ws.append(["Theme", theme[0], theme[1], "-"])
        for english, chinese, example in patterns:
            ws.append(["Sentence Pattern", english, chinese, example])
        for english, chinese, example in words:
            ws.append(["Vocabulary", english, chinese, example])
        ws.append([])

    wb.save(SRC)
    print(f"已寫入 Book 1：{len(UNITS)} 個單元、"
          f"{sum(len(u[4]) for u in UNITS)} 個句型、{sum(len(u[5]) for u in UNITS)} 個單字")


if __name__ == "__main__":
    main()
